#!/usr/bin/env python3
"""
Safespill Hangar Projects Scraper
Main execution script for collecting MRO hangar construction and retrofit projects
"""

import os
import sys
import logging
from datetime import datetime
from typing import List, Dict, Any

# Add current directory to path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from news_scraper import NewsScraper
from excel_handler import ExcelHandler
from email_sender import EmailSender
from scheduler import ScrapingScheduler
from config import Config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class SafespillScraper:
    def __init__(self):
        self.config = Config()
        self.regions = ['uk_na', 'emea']
        self.existing_urls = self._load_existing_urls()
        
    def _load_existing_urls(self) -> set:
        """Load all existing URLs from both region Excel files into a set"""
        urls = set()
        for region in self.regions:
            handler = ExcelHandler(region)
            try:
                if not os.path.exists(handler.filepath):
                    continue
                from openpyxl import load_workbook
                wb = load_workbook(handler.filepath)
                ws = wb.active
                url_column = self.config.EXCEL_FIELDS.index('Source URL') + 1
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= url_column:
                        url = row[url_column - 1]
                        if url:
                            urls.add(url)
            except Exception as e:
                print(f"Error loading URLs from {handler.filepath}: {str(e)}")
        return urls

    def filter_new_urls(self, articles: list) -> list:
        """Filter out articles whose 'Source URL' is already in self.existing_urls. Update self.existing_urls with new ones."""
        new_articles = []
        for article in articles:
            url = article.get('Source URL', '')
            if url and url not in self.existing_urls:
                new_articles.append(article)
                self.existing_urls.add(url)
        return new_articles

    def run_scraping_for_region(self, region: str, is_backfill: bool = False) -> bool:
        """Run the complete scraping process for a specific region"""
        try:
            logger.info(f"Starting scraping for region: {region}")
            
            # Initialize components
            scraper = NewsScraper(region)
            excel_handler = ExcelHandler(region)
            email_sender = EmailSender()
            
            # Scrape news articles
            logger.info(f"Scraping news articles for {region}...")
            articles = scraper.scrape_all_sources(is_backfill=is_backfill)
            
            if not articles:
                logger.warning(f"No articles found for region: {region}")
                return False
            
            # Process articles
            logger.info(f"Processing {len(articles)} articles...")
            processed_articles = scraper.process_articles(articles)
            
            # Filter out existing articles using self.existing_urls
            new_articles = self.filter_new_urls(processed_articles)
            
            if not new_articles:
                logger.info(f"No new articles found for region: {region}")
                return True
            
            # Update Excel file
            logger.info(f"Updating Excel file with {len(new_articles)} new articles...")
            excel_filepath = excel_handler.update_existing_excel(new_articles)
            
            # Send email report
            logger.info("Sending email report...")
            email_success = email_sender.send_report_email(
                excel_filepath, region, len(new_articles)
            )
            
            # Log summary
            file_info = excel_handler.get_file_info()
            logger.info(f"Scraping completed for {region}:")
            logger.info(f"  - New articles: {len(new_articles)}")
            logger.info(f"  - Total articles in file: {file_info['row_count']}")
            logger.info(f"  - Email sent: {email_success}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error in scraping for region {region}: {str(e)}")
            return False
    
    def run_backfill(self) -> bool:
        """Run initial backfill for past 12 months"""
        logger.info("Starting initial backfill for past 12 months...")
        
        success = True
        for region in self.regions:
            logger.info(f"Running backfill for region: {region}")
            if not self.run_scraping_for_region(region, is_backfill=True):
                success = False
        
        if success:
            logger.info("Backfill completed successfully")
        else:
            logger.error("Backfill completed with errors")
        
        return success
    
    def run_weekly_scraping(self) -> bool:
        """Run weekly scraping for all regions"""
        logger.info("Starting weekly scraping...")
        
        success = True
        for region in self.regions:
            logger.info(f"Running weekly scraping for region: {region}")
            if not self.run_scraping_for_region(region, is_backfill=False):
                success = False
        
        if success:
            logger.info("Weekly scraping completed successfully")
        else:
            logger.error("Weekly scraping completed with errors")
        
        return success
    
    def test_configuration(self) -> bool:
        """Test all configuration settings"""
        logger.info("Testing configuration...")
        
        # Test API key
        if not self.config.SERPAPI_KEY:
            logger.error("SERPAPI_KEY not configured")
            return False
        
        # Test email configuration
        email_sender = EmailSender()
        email_test = email_sender.test_email_configuration()
        
        # Test file paths
        for region in self.regions:
            excel_handler = ExcelHandler(region)
            file_info = excel_handler.get_file_info()
            logger.info(f"Excel file for {region}: {file_info['filepath']}")
        
        logger.info("Configuration test completed")
        return email_test

def main():
    """Main execution function"""
    scraper = SafespillScraper()
    
    # Check command line arguments
    if len(sys.argv) > 1:
        command = sys.argv[1].lower()
        
        if command == 'backfill':
            # Run initial backfill
            scraper.run_backfill()
            
        elif command == 'weekly':
            # Run weekly scraping
            scraper.run_weekly_scraping()
            
        elif command == 'test':
            # Test configuration
            scraper.test_configuration()
            
        elif command == 'schedule':
            # Start scheduler
            scheduler = ScrapingScheduler()
            scheduler.schedule_weekly_run(scraper.run_weekly_scraping)
            scheduler.run_scheduler()
            
        elif command == 'region':
            # Run for specific region
            if len(sys.argv) > 2:
                region = sys.argv[2].lower()
                if region in ['uk_na', 'emea']:
                    scraper.run_scraping_for_region(region)
                else:
                    logger.error(f"Invalid region: {region}. Use 'uk_na' or 'emea'")
            else:
                logger.error("Please specify a region: 'uk_na' or 'emea'")
        elif command == 'email':
            # Send test email
            email_sender = EmailSender()
            if email_sender.test_send():
                logger.info("Test email sent successfully")
            else:
                logger.error("Failed to send test email")
        else:
            print("Usage:")
            print("  python main.py backfill    - Run initial 12-month backfill")
            print("  python main.py weekly      - Run weekly scraping")
            print("  python main.py test        - Test configuration")
            print("  python main.py schedule    - Start scheduler")
            print("  python main.py region uk_na - Run for specific region")
            print("  python main.py region emea  - Run for specific region")
    
    else:
        # Default: run weekly scraping
        logger.info("No command specified, running weekly scraping...")
        scraper.run_weekly_scraping()

if __name__ == "__main__":
    main() 