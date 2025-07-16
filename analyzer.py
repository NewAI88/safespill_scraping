import openai
import json
import time
from typing import List, Dict, Any
import math
from datetime import datetime

class HangarArticleAnalyzer:
    def __init__(self, api_key: str, model: str = "gpt-4o", default_region: str = "UK_NA"):
        self.client = openai.OpenAI(api_key=api_key)
        self.model = model
        self.batch_size = 10  # Safe batch size for 50-token articles
        self.default_region = default_region
        self.current_date = datetime.now().strftime("%B %d, %Y")

    def create_batch_prompt(self, articles: List[Dict[str, str]]) -> str:
        """Create optimized batch prompt for multiple articles"""
        
        system_prompt = f"""Analyze aircraft hangar articles and return structured JSON.

IMPORTANT: Today's date is {self.current_date}. Use this for determining completion status.

DETAILED FIELD EXPLANATIONS:

1. **is_hangar_related** (boolean):
   - True ONLY if article is directly related to:
     * Aircraft hangar construction (new builds)
     * Aircraft hangar renovation/expansion/demolition
     * Aircraft maintenance facilities (MRO - Maintenance, Repair, Overhaul)
     * Hangar equipment, doors, systems, infrastructure
     * Hangar operations, leasing, management,
     * Airport infrastructure projects that INCLUDE hangar facilities
     * Maintenance base construction or expansion
     * Flight training facilities/academies with hangar components
     * Aviation training bases and facilities
   - False for:
     * General aviation news
     * Aircraft purchases/sales
     * Flight operations and schedules
     * Airport terminals, runways ONLY (without hangar components)
     * Airline business news
     * Aircraft repairs happening IN a hangar (focus on repair, not hangar facility)
     * Emergency aircraft situations that happen to mention hangar location

2. **country** (string):
   - 2-letter ISO country code (US, CA, UK, FR, DE, AU, JP, etc.)
   - Use "N/A" if country cannot be determined from title/summary
   
3. **region** (string):
   - Must be exactly ONE of these: "UK_NA", "EMEA", "APAC", "LATAM"
   - UK_NA: United Kingdom, United States, Canada
   - EMEA: Europe (except UK), Middle East, Africa
   - APAC: Asia-Pacific (including Australia, New Zealand)
   - LATAM: Latin America (Mexico, Central America, South America)

4. **completion_status** (boolean):
   - CRITICAL: Consider today's date is {self.current_date}
   - True if:
     * Project is already completed/finished/opened ("opens", "completed", "inaugurated")
     * Project had a deadline that has already passed (e.g., "by end of 2024", "by Q1 2025", "by June 2025")
     * Project was scheduled for a past date (e.g., "opens in 2024", "scheduled for early 2025")
   - False if:
     * Project is planned for future dates ("will open in 2026", "planned for 2027")
     * Project is under construction with no specific completion date
     * Project is in planning phase

For each article, return:
{
    "article_id": number,
    "is_hangar_related": boolean,
    "country": string,
    "region": string,
    "completion_status": boolean
}

Articles to analyze:
"""
        
        articles_text = ""
        for i, article in enumerate(articles, 1):
            articles_text += f"\n{i}. Title: {article['Project Title']}\nSummary: {article['Summary']}\n"
        
        response_format = f"""
Return JSON array:
{{
    "results": [
        // {len(articles)} objects with article_id 1-{len(articles)}
        // Each object must have: article_id, is_hangar_related, region, country, completion_status
    ]
}}"""
        
        return system_prompt + articles_text + response_format

    def validate_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Validate and clean result fields"""
        
        # Get country first
        country = result.get('country', 'N/A')
        if country and len(country) == 2 and country != 'N/A':
            country = country.upper()
            result['country'] = country
        else:
            result['country'] = 'N/A'
            country = 'N/A'
        
        # Auto-derive region from country if not provided or invalid
        valid_regions = ["UK_NA", "EMEA", "APAC", "LATAM"]
        current_region = result.get('region')
        
        if current_region not in valid_regions:
            # Auto-derive region from country
            if country in ['US', 'CA', 'UK', 'GB']:
                result['region'] = 'UK_NA'
            elif country in ['FR', 'DE', 'IT', 'ES', 'NL', 'BE', 'CH', 'AT', 'SE', 'NO', 'DK', 'FI', 'PL', 'CZ', 'HU', 'PT', 'IE', 'GR', 'RO', 'BG', 'HR', 'SI', 'SK', 'LT', 'LV', 'EE', 'MT', 'CY', 'LU', 'AE', 'SA', 'QA', 'KW', 'BH', 'OM', 'JO', 'LB', 'IL', 'TR', 'EG', 'ZA', 'MA', 'TN', 'DZ', 'LY', 'SD', 'ET', 'KE', 'UG', 'TZ', 'ZM', 'ZW', 'BW', 'NA', 'MZ', 'AO', 'GH', 'NG', 'CI', 'SN', 'ML', 'BF', 'NE', 'TD', 'CM', 'CF', 'CG', 'CD', 'GA', 'GQ', 'ST', 'CV', 'GM', 'GW', 'SL', 'LR', 'GN', 'BJ', 'TG', 'RW', 'BI', 'DJ', 'SO', 'ER', 'MW', 'MG', 'MU', 'SC', 'KM', 'YT', 'RE']:
                result['region'] = 'EMEA'
            elif country in ['JP', 'CN', 'KR', 'SG', 'TH', 'MY', 'ID', 'PH', 'VN', 'IN', 'PK', 'BD', 'LK', 'MM', 'KH', 'LA', 'BN', 'TL', 'MN', 'KZ', 'KG', 'TJ', 'TM', 'UZ', 'AF', 'IR', 'IQ', 'SY', 'YE', 'PS', 'AU', 'NZ', 'PG', 'FJ', 'SB', 'NC', 'PF', 'WS', 'VU', 'TO', 'TV', 'NR', 'KI', 'PW', 'MH', 'FM', 'GU', 'MP', 'AS', 'CK', 'NU', 'TK', 'WF']:
                result['region'] = 'APAC'
            elif country in ['MX', 'BR', 'AR', 'CL', 'CO', 'PE', 'VE', 'EC', 'BO', 'PY', 'UY', 'GY', 'SR', 'GF', 'CR', 'PA', 'NI', 'HN', 'GT', 'BZ', 'SV', 'CU', 'DO', 'HT', 'JM', 'TT', 'BB', 'GD', 'VC', 'LC', 'AG', 'DM', 'KN', 'BS', 'BM', 'PR', 'VI', 'AW', 'CW', 'SX', 'BQ', 'MQ', 'GP', 'BL', 'MF', 'PM', 'TC', 'AI', 'VG', 'KY', 'MS', 'FK']:
                result['region'] = 'LATAM'
            else:
                result['region'] = 'N/A' # self.default_region  # Default fallback
        else:
            result['region'] = current_region

        # Ensure boolean fields
        result['is_hangar_related'] = bool(result.get('is_hangar_related', True))
        result['completion_status'] = bool(result.get('completion_status', False))
        
        return result

    def analyze_batch(self, articles: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        """Analyze a batch of articles"""
        
        try:
            prompt = self.create_batch_prompt(articles)
            
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0.1,
                # max_tokens=len(articles) * 200  # Increased for detailed responses
            )
            
            result = json.loads(response.choices[0].message.content)
            raw_results = result.get("results", [])
            
            # Validate each result
            validated_results = []
            for idx, raw_result in enumerate(raw_results):
                validated_result = self.validate_result(raw_result)
                validated_results.append(articles[idx] | {
                    "Country": validated_result.get("country"),
                    "Region": validated_result.get("region"),
                    "Is Hangar Related": validated_result.get("is_hangar_related"),
                    "Completion Status": validated_result.get("completion_status")
                })  # Merge original article data with analysis result

            return validated_results
            
        except Exception as e:
            print(f"Error in batch analysis: {e}")
            return []

    def process_all_articles(self, articles: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        """Process all articles in batches"""
        
        all_results = []
        total_batches = math.ceil(len(articles) / self.batch_size)
        
        for i in range(0, len(articles), self.batch_size):
            batch = articles[i:i + self.batch_size]
            batch_num = (i // self.batch_size) + 1
            
            print(f"Processing batch {batch_num}/{total_batches} ({len(batch)} articles)")
            
            batch_results = self.analyze_batch(batch)
            
            if batch_results:
                all_results.extend(batch_results)
            
            # Rate limiting - be respectful to API
            time.sleep(1)
        
        filtered_results = [r for r in all_results if r.get("Is Hangar Related", True) and r.get("Completion Status", False) == False and r.get("Region") == self.default_region]

        return filtered_results

    def generate_summary_report(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate detailed summary report"""
        
        total_articles = len(results)
        hangar_related = [r for r in results if r.get('is_hangar_related', False)]
        completed_projects = [r for r in hangar_related if r.get('completion_status', False)]
        
        # Region breakdown
        region_counts = {}
        for result in hangar_related:
            region = result.get('region', 'N/A')
            region_counts[region] = region_counts.get(region, 0) + 1
        
        # Country breakdown
        country_counts = {}
        for result in hangar_related:
            country = result.get('country', 'N/A')
            country_counts[country] = country_counts.get(country, 0) + 1
                
        return {
            'summary': {
                'total_articles': total_articles,
                'hangar_related': len(hangar_related),
                'not_hangar_related': total_articles - len(hangar_related),
                'completed_projects': len(completed_projects),
                'in_progress_projects': len(hangar_related) - len(completed_projects)
            },
            'region_breakdown': region_counts,
            'country_breakdown': country_counts,
            'sample_results': results[:5]  # First 5 results as examples
        }

# Usage example
def main():
    # Sample articles (replace with your data)
    articles = [
        # 1
        {
            'title': 'Lincoln Airport gets welcome surprises with runway project, added flights',
            'summary': 'The airport received bids, several under initial estimates, for one of the largest projects in its history. More good news: United is prepared to add more flights out of Lincoln.'
        },
        # 2
        {
            'title': 'Flydubai Breaks Ground On MRO Facility At Dubai South',
            'summary': 'Flydubai said construction is due to be completed in the last quarter of 2026. The maintenance center at DWC coincides with ...'
        },
        # 3
        {
            'title': 'Bombardier Inaugurates Expanded London Biggin Hill Service Centre, the Largest Business Jet Maintenance Repair and Overhaul (MRO) Facility in the UK',
            'summary': ''
        },
        # 4
        {
            'title': "UK’s stranded F-35B fighter jet moved to Air India hangar in Kerala as British team arrives for repairs",
            'summary': ''
        },
        # 5
        {
            'title': 'Furore over allocation of land for aircraft hangar',
            'summary': ''
        },
        # 6
        {
            'title': 'Luxaviation Unveils New Belgium Hangar',
            'summary': ''
        },
        # 7
        {
            'title': 'Ledcor hosts CAWIC on Hamilton hangar project tour',
            'summary': ''
        },
        # 8
        {
            'title': 'Rs 50-crore project aims to turn Kochi into premier aircraft maintenance hub',
            'summary': ''
        },
        # 9
        {
            'title': 'Sanad expands MRO services for CFM LEAP engines',
            'summary': 'The move, in partnership with GE Aerospace and Safran Aircraft Engines, enables full CFM LEAP engine overhaul and test ...'
        }
        # ... add your articles
    ]
    
    import os
    import dotenv

    dotenv.load_dotenv()

    analyzer = HangarArticleAnalyzer(api_key=os.getenv("OPENAI_API_KEY"))
    
    # Process all articles
    results = analyzer.process_all_articles(articles)
    
    # Generate detailed report
    report = analyzer.generate_summary_report(results)
    
    print(json.dumps(report, indent=2))

def filter_files():
    import os
    import glob
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    reports_folder = "reports"
    output_folder = "reports"
    default_region="UK_NA"
    os.makedirs(output_folder, exist_ok=True)

    excel_files = glob.glob(os.path.join(reports_folder, "*.xlsx"))


    # Define fills
    red_fill = PatternFill(start_color="fa0228", end_color="fa0228", fill_type="solid")
    blue_fill = PatternFill(start_color="00f7ff", end_color="00f7ff", fill_type="solid")
    yellow_fill = PatternFill(start_color="ffe100", end_color="ffe100", fill_type="solid")

    for file in excel_files:
        file_name = os.path.basename(file)
        default_region = file_name[:-21] if file_name.endswith("_hangar_projects.xlsx") else default_region
        analyzer = HangarArticleAnalyzer(api_key=os.getenv("OPENAI_API_KEY"), default_region=default_region)
        df = pd.read_excel(file)
        articles = []
        for idx, row in df.iterrows():
            articles.append({
                "title": str(row.get("Project Title", "")),
                "summary": str(row.get("Summary", ""))
            })

        results = analyzer.process_all_articles(articles)

        # Add analysis columns
        df["is_hangar_related"] = [r.get("is_hangar_related", False) for r in results]
        df["region"] = [r.get("region", "N/A") for r in results]
        df["country"] = [r.get("country", "N/A") for r in results]
        df["completion_status"] = [r.get("completion_status", False) for r in results]

        # Save to new Excel file
        output_path = os.path.join(output_folder, os.path.basename(file))
        df.to_excel(output_path, index=False)

        # Apply background colors
        wb = load_workbook(output_path)
        ws = wb.active
        for i, r in enumerate(results, start=2):  # Excel rows start at 2 (header is 1)
            # 1. Red for outside of territory (region == "N/A")
            if r.get("region") != default_region:
                ws[f"A{i}"].fill = red_fill
                continue
            # 2. Yellow for not hangar project
            elif not r.get("is_hangar_related", False):
                # for cell in ws[i]:
                ws[f"A{i}"].fill = yellow_fill
                continue
            # 3. Blue for completed articles
            elif r.get("completion_status", False):
                ws[f"A{i}"].fill = blue_fill

        wb.save(output_path)
    pass

if __name__ == "__main__":
    filter_files()