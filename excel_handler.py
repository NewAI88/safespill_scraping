import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config import Config

class ExcelHandler:
    def __init__(self, region: str):
        self.region = region
        self.config = Config()
        self.reports_dir = self.config.REPORTS_DIR
        
        # Ensure reports directory exists
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Define filename based on region
        if region == 'uk_na':
            self.filename = 'UK_NA.xlsx'
        elif region == 'emea':
            self.filename = 'EMEA.xlsx'
        else:
            self.filename = f'{region}.xlsx'

        self.filepath = os.path.join(self.reports_dir, self.filename)
        self.pattern_set = self._load_all_patterns()

    def _load_all_patterns(self) -> set:
        """Load all patterns from every tab in the Excel file."""
        pattern_set = set()
        if not os.path.exists(self.filepath):
            return pattern_set
        try:
            wb = load_workbook(self.filepath)
            headers = self.config.EXCEL_FIELDS
            pattern_idx = None
            if 'patterns' in headers:
                pattern_idx = headers.index('patterns')
            else:
                return pattern_set
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > pattern_idx:
                        patterns = row[pattern_idx]
                        if patterns:
                            if isinstance(patterns, str):
                                for p in patterns.split(','):
                                    p = p.strip()
                                    if p:
                                        pattern_set.add(p)
                            elif isinstance(patterns, list):
                                for p in patterns:
                                    if p:
                                        pattern_set.add(p)
            return pattern_set
        except Exception as e:
            print(f"Error loading patterns: {str(e)}")
            return pattern_set
    
    def create_new_excel(self) -> str:
        """Create a new Excel file with the scraped data"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Hangar Projects"
            
            # Add headers
            headers = self.config.EXCEL_FIELDS
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # # Add data rows
            # for row_idx, row_data in enumerate(data, 2):
            #     for col_idx, field in enumerate(headers, 1):
            #         value = row_data.get(field, '')
            #         cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
            #         # Add hyperlink for Source URL column
            #         if field == 'Source URL' and value:
            #             cell.hyperlink = value
            #             cell.font = Font(color="0000FF", underline="single")
            #             cell.value = value
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(self.filepath)
            
            print(f"Created new Excel file: {self.filepath}")
            return self.filepath
            
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            raise
    
    def update_existing_excel(self, new_data: List[Dict[str, Any]]) -> tuple[str, int]:
        """Update existing Excel file by creating a new tab with week number, filtering by pattern_set"""
        max_retries = 3
        retry_delay = 2  # seconds
        
        for attempt in range(max_retries):
            try:
                if not os.path.exists(self.filepath):
                    self.create_new_excel()
                
                # Load existing workbook
                wb = load_workbook(self.filepath)
                
                # Get current week number
                current_week = datetime.now().isocalendar()[1]
                new_tab_name = f"W{current_week}"
                
                # Create new worksheet
                ws = wb.create_sheet(title=new_tab_name)
                
                # Get headers
                headers = self.config.EXCEL_FIELDS
                
                # Add headers to new worksheet
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                
                # Filter new data to avoid duplicates (by pattern)
                filtered_new_data = []
                for article in new_data:
                    patterns = article.get('patterns', [])
                    if isinstance(patterns, str):
                        patterns = [p.strip() for p in patterns.split(',') if p.strip()]
                    is_new = any(p not in self.pattern_set for p in patterns)
                    if is_new:
                        filtered_new_data.append(article)
                        for p in patterns:
                            self.pattern_set.add(p)
                
                # Add filtered new data to the new worksheet
                for row_idx, row_data in enumerate(filtered_new_data, 2):
                    for col_idx, field in enumerate(headers, 1):
                        value = row_data.get(field, '')
                        if isinstance(value, list):
                            value = ', '.join(value)
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # Add hyperlink for Source URL column
                        if field == 'Source URL' and value:
                            cell.hyperlink = value
                            cell.font = Font(color="0000FF", underline="single")
                            cell.value = value
                
                # Auto-adjust column widths for new worksheet
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save the workbook
                wb.save(self.filepath)
                
                print(f"Created new tab '{new_tab_name}' with {len(filtered_new_data)} articles in Excel file: {self.filepath}")
                return self.filepath, len(filtered_new_data)
                
            except PermissionError as e:
                if attempt < max_retries - 1:
                    print(f"Permission denied on attempt {attempt + 1}. File may be open in another application.")
                    print(f"Please close the Excel file '{self.filename}' and retrying in {retry_delay} seconds...")
                    import time
                    time.sleep(retry_delay)
                    retry_delay *= 2  # Exponential backoff
                else:
                    print(f"Error updating Excel file after {max_retries} attempts: {str(e)}")
                    print(f"Please ensure the file '{self.filename}' is not open in Excel or another application.")
                    raise
            except Exception as e:
                print(f"Error updating Excel file: {str(e)}")
                raise
    
    def get_existing_urls(self) -> set:
        """Get existing URLs from all tabs in the Excel file to avoid duplicates"""
        try:
            if not os.path.exists(self.filepath):
                return set()
            
            wb = load_workbook(self.filepath)
            existing_urls = set()
            url_column = self.config.EXCEL_FIELDS.index('Source URL') + 1
            
            # Iterate through all worksheets
            for ws in wb.worksheets:
                # Read URLs from existing data (skip header row)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= url_column:
                        url = row[url_column - 1]
                        if url:
                            existing_urls.add(url)
            
            return existing_urls
            
        except Exception as e:
            print(f"Error reading existing URLs: {str(e)}")
            return set()
    
    def filter_new_articles(self, articles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Filter out articles that already exist in the Excel file"""
        existing_urls = self.get_existing_urls()
        new_articles = []
        
        for article in articles:
            url = article.get('Source URL', '')
            if url and url not in existing_urls:
                new_articles.append(article)
        
        print(f"Found {len(new_articles)} new articles out of {len(articles)} total")
        return new_articles
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get information about the Excel file"""
        try:
            if not os.path.exists(self.filepath):
                return {
                    'exists': False,
                    'filepath': self.filepath,
                    'size': 0,
                    'last_modified': None,
                    'row_count': 0
                }
            
            stat = os.stat(self.filepath)
            
            # Count rows in Excel file
            wb = load_workbook(self.filepath)
            ws = wb.active
            row_count = ws.max_row - 1  # Subtract header row
            
            return {
                'exists': True,
                'filepath': self.filepath,
                'size': stat.st_size,
                'last_modified': datetime.fromtimestamp(stat.st_mtime),
                'row_count': max(0, row_count)
            }
            
        except Exception as e:
            print(f"Error getting file info: {str(e)}")
            return {
                'exists': False,
                'filepath': self.filepath,
                'size': 0,
                'last_modified': None,
                'row_count': 0
            } 